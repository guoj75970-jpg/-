import pandas as pd, numpy as np, math, json, shutil, re
from pathlib import Path
from collections import defaultdict, Counter
from copy import copy
from openpyxl import load_workbook

DAYS=[28]
BASE=Path(r'D:\HY-日度数据监测\HY-日志导出月度汇总\7月视频审核操作日志')
XLSX=Path(r'D:\虎牙数据汇总\虎牙结算数据\26年结算\7月结算\7月人审量.xlsx')
ROOT=Path(r'C:\Users\guojing17\.easyclaw\workspace\虎牙数据看板')
HTML=ROOT/'index.html'
BACKUP=XLSX.with_name('7月人审量_追加7.28前备份.xlsx')
HTML_BACKUP=ROOT/'index_追加7.28前备份.html'
COLS=['vid','所属应用','视频时长','上传时间','工单生成时间','进入审核时间','审核完成时间','审核结果','审核人']
VALID_APPS=['虎牙视频','虎牙直播回放切片视频','虎牙AI礼物特效视频','虎牙一起看视频']

def dursec(s):
    try:
        p=[float(x) for x in str(s).split(':')]
        return p[0]*3600+p[1]*60+p[2] if len(p)==3 else np.nan
    except: return np.nan

def norm_type(s): return s.fillna('').astype(str).str.replace('—','-',regex=False).str.replace('–','-',regex=False).str.strip()
def enrich(d):
    d=d.copy(); d['duration_sec']=d['视频时长'].map(dursec)
    for c in ['工单生成时间','进入审核时间','审核完成时间']: d[c]=pd.to_datetime(d[c],errors='coerce')
    d['queue_min']=(d['进入审核时间']-d['工单生成时间']).dt.total_seconds()/60
    d['queue_excluded']=norm_type(d['视频类型']).eq('爬取视频-重新推审')
    d['audit_sec']=(d['审核完成时间']-d['进入审核时间']).dt.total_seconds()
    d['mult']=d['audit_sec']/d['duration_sec']
    d.loc[(d['duration_sec']<=0)|(d['audit_sec']<0)|(d['audit_sec']>3600),'mult']=np.nan
    d['weight']=np.where(d['duration_sec']>3600,1.5,0.78)
    d['bad']=d['审核结果'].isin(['不通过','已删除'])
    return d

def metrics(d):
    q=d.loc[(~d.queue_excluded)&d.queue_min.between(0,60),'queue_min']
    return {'actual':len(d),'trans':math.floor(d.weight.sum()),'passed':int((d['审核结果']=='通过').sum()),'risk':int((d['审核结果']=='通过（风险）').sum()),'bad':int(d.bad.sum()),'viol':round(d.bad.sum()/len(d)*100,2),'queue':round(q.mean(),2),'queue_n':int(q.count()),'queue_sum':float(q.sum()),'mult':round(d.mult.mean(),2)}

def load_day(day):
    p=BASE/f'视频审核操作日志-7.{day}.csv'; assert p.exists(),p
    raw=pd.read_csv(p,encoding='utf-8',low_memory=False,dtype={'vid':str})
    miss=[c for c in COLS+['视频类型'] if c not in raw.columns]; assert not miss,miss
    b=raw[raw['审核人'].fillna('').astype(str).str.contains('易盾BPO')].copy(); assert len(b)>0,f'7.{day} no 易盾BPO data'
    b['vid']=b['vid'].fillna('').astype(str).str.replace('\t','',regex=False).str.strip(); assert b['vid'].ne('').all()
    return raw,enrich(b)

def make_report(day,b,m):
    m=m.copy(); m['people']=int(b['审核人'].nunique()); m['percap']=math.floor(m['trans']/m['people'])
    persons=[]
    for reviewer,g in b.groupby('审核人'):
        x=metrics(g); persons.append({'name':reviewer.replace('-易盾BPO',''),'audit':x['actual'],'trans':x['trans'],'passed':x['passed'],'bad':x['bad'],'viol':x['viol'],'queue':x['queue'],'mult':x['mult']})
    persons.sort(key=lambda x:(-x['audit'],x['name']))
    apps={str(k):int(v) for k,v in b['所属应用'].value_counts().to_dict().items()}; other={k:v for k,v in apps.items() if k not in VALID_APPS}; high=[p['name'] for p in persons if p['viol']>15]
    cap='🟢' if m['percap']>=230 else '🔴'; vs='🟢' if not high else '🟡'; qs='🟢' if m['queue']<=2 else '🔴'; iso=f'2026-07-{day:02d}'
    L=[f'📊 【虎牙视频审核日报】{iso}','','📈 整体业务情况','指标│数值',f"实际审核总量│{m['actual']} 条",f"转化后审核总量│{m['trans']} 条",f"通过量│{m['passed']} 条",f"通过（风险）量│{m['risk']} 条",f"不通过量│{m['bad']} 条",f"违规率│{m['viol']:.2f}%",f"平均排队时长│{m['queue']:.2f} 分钟",f"平均审核倍率│{m['mult']:.2f}",f"当班次人数│{m['people']} 人",f"人均转化后审核量│{m['percap']} 条/人",'','👥 当班次人员情况','姓名│审核量│转化后│通过量│不通过量│违规率│排队时长│审核倍率']
    for p in persons:
        qv=f"{p['queue']:.2f}" if pd.notna(p['queue']) else '无有效数据'; mv=f"{p['mult']:.2f}" if pd.notna(p['mult']) else '无有效数据'
        L.append(f"{p['name']}│{p['audit']}│{p['trans']}│{p['passed']}│{p['bad']}│{p['viol']:.2f}%│{qv}│{mv}")
    L += ['', '⚠️ 异常预警','预警类型│状态│详情',f"产能异常│{cap}│人均转化后审核量 {m['percap']} 条/人 {'≥' if m['percap']>=230 else '<'} 目标230",f"违规率异常│{vs}│"+('个人违规率均≤15%' if not high else '超过15%：'+'、'.join(high)),f"排队时长异常│{qs}│平均排队时长 {m['queue']:.2f} 分钟 {'≤' if m['queue']<=2 else '>'} 目标2分钟",'','💡 建议措施',('- 产能提升：当前人效达标，继续保持班次节奏。' if m['percap']>=230 else f"- 产能提升：人均转化后仅 {m['percap']} 条，建议复盘低产时段并优化排班。"),('- 违规率控制：个人违规率均在正常范围内。' if not high else '- 违规率控制：重点复核 '+'、'.join(high)+' 的高违规样本，确认业务分配与规则执行情况。'),('- 排队时长关注：均值超过2分钟目标，建议排查高峰时段积压。' if m['queue']>2 else '- 排队时长关注：当前正常，继续监控高峰时段。'),'','📊 通道分布','通道名称│数量│状态']
    for a in VALID_APPS: L.append(f"{a}│{apps.get(a,0)}│"+('✅' if apps.get(a,0) else 'ℹ️ 当日无数据'))
    L.append(f"其他异常通道│{sum(other.values())}│"+('⚠️ '+json.dumps(other,ensure_ascii=False) if other else '✅ 无'))
    report=ROOT/f'虎牙视频审核日报-{iso}.txt'; report.write_text('\n'.join(L),encoding='utf-8')
    summary={'day':f'7.{day}','metrics':m,'persons':persons,'apps':apps,'other_apps':other,'queue_excluded_records':int(b.queue_excluded.sum()),'report':str(report)}
    (ROOT/f'analysis_7.{day}.json').write_text(json.dumps(summary,ensure_ascii=False,indent=2),encoding='utf-8')
    return summary

# Read new source files once and create reports.
loaded={}; reports={}
for day in DAYS:
    raw,b=load_day(day); loaded[day]=(raw,b); reports[day]=make_report(day,b,metrics(b))

# Append each missing date to monthly settlement workbook, preserving every source record.
old=pd.read_excel(XLSX,dtype={'vid':str,'数据日期':str}); assert list(old.columns[:10])==COLS+['数据日期']
if not BACKUP.exists(): shutil.copy2(XLSX,BACKUP)
wb=load_workbook(XLSX); ws=wb[wb.sheetnames[0]]; last=ws.max_row
existing=old['数据日期'].astype(str).str.replace(r'\.0$','',regex=True)
added={}
for day in DAYS:
    b=loaded[day][1]; ds=f'7.{day}'
    if (existing==ds).any(): added[day]=0; continue
    for row in b[COLS].itertuples(index=False,name=None):
        last+=1; vals=list(row)+[ds]
        for ci,val in enumerate(vals,1):
            cell=ws.cell(last,ci); cell.value=None if pd.isna(val) else val; src=ws.cell(last-1,ci)
            if src.has_style: cell._style=copy(src._style)
            if ci in (5,6,7): cell.number_format='yyyy-mm-dd hh:mm:ss'
    added[day]=len(b)
wb.save(XLSX)
post=pd.read_excel(XLSX,dtype={'vid':str,'数据日期':str}); pdts=post['数据日期'].astype(str).str.replace(r'\.0$','',regex=True)
for day in DAYS:
    b=loaded[day][1]; chk=post.loc[pdts==f'7.{day}']
    assert len(chk)==len(b),(day,len(chk),len(b)); assert chk['vid'].fillna('').astype(str).str.strip().ne('').all()
assert len(post)-len(old)==sum(added.values())

# Rebuild dashboard from all source CSVs through 7.28.
labels=[]; audit=[]; trans=[]; queues=[]; viol=[]; zv=[]; zq=[]; people_total=0; bad_total=0; qsum=0.; qcount=0.; pw=defaultdict(float); pdays=defaultdict(set); hourly=Counter(); apps=Counter()
for day in range(1,29):
    raw,b=loaded[day] if day in loaded else load_day(day); mm=metrics(b)
    labels.append(f'7.{day}'); audit.append(mm['actual']); trans.append(mm['trans']); queues.append(mm['queue']); viol.append(mm['viol']); people_total+=b['审核人'].nunique(); bad_total+=mm['bad']; qsum+=mm['queue_sum']; qcount+=mm['queue_n']; hourly.update(b['审核完成时间'].dt.hour.dropna().astype(int).tolist()); apps.update(b['所属应用'].dropna().tolist())
    for name,g in b.groupby('审核人'): pw[name.replace('-易盾BPO','')]+=g.weight.sum(); pdays[name.replace('-易盾BPO','')].add(day)
    zr=raw[raw['审核人'].fillna('').astype(str).str.contains('证通BPO')].copy(); z=enrich(zr); zm=metrics(z) if len(z) else {'actual':0,'queue':np.nan}; zv.append(zm['actual']); zq.append(zm['queue'])
pr=sorted(((n,round(pw[n]/len(pdays[n]),1)) for n in pw),key=lambda x:(-x[1],x[0]))[:12]
cum_actual=sum(audit); cum_trans=sum(trans); eff=math.floor(cum_trans/people_total); cum_viol=round(bad_total/cum_actual*100,2); cum_queue=round(qsum/qcount,2); hr=[hourly[i] for i in range(24)]
if not HTML_BACKUP.exists(): shutil.copy2(HTML,HTML_BACKUP)
text=HTML.read_text(encoding='utf-8')
def sub(p,r):
    global text
    text,n=re.subn(p,r,text,count=1,flags=re.S); assert n==1,(p,n)
def arr(x): return json.dumps(x,ensure_ascii=False,separators=(',',':'))
sub(r'(<div class="label">累计实际审核量</div>\s*<div class="value blue">)[^<]+',rf'\g<1>{cum_actual:,} ')
sub(r'(<div class="label">累计转化后审核量</div>\s*<div class="value green">)[^<]+',rf'\g<1>{cum_trans:,} ')
sub(r'(<div class="label">日均人效\(转化后\)</div>\s*<div class="value purple">)[^<]+',rf'\g<1>{eff} ')
sub(r'(<div class="label">平均违规率</div>\s*<div class="value orange">)[^<]+',rf'\g<1>{cum_viol:.2f} ')
sub(r'(<div class="label">平均排队时长</div>\s*<div class="value red">)[^<]+',rf'\g<1>{cum_queue:.2f} ')
sub(r'7月在班总人力\d+人',f'7月在班总人力{people_total}人')
for var,val in [('dates',labels),('auditVolumes',audit),('transformedVolumes',trans),('queueTimes',queues),('personNames',[x[0] for x in pr]),('personValues',[x[1] for x in pr]),('hourlyData',hr),('julyLabels',labels),('yidunVolumes',audit),('zhengtongVolumes',zv),('yidunQueueTimes',queues),('zhengtongQueueTimes',zq)]: sub(rf'const {var} = \[[^;]*\];',f'const {var} = {arr(val)};')
sub(r"(labels: )\[[^\]]*\](,\s*datasets: \[\{\s*label: '每日违规率')",rf'\g<1>{arr(labels)}\g<2>')
sub(r"(label: '每日违规率',\s*data: )\[[^\]]*\]",rf'\g<1>{arr(viol)}')
sub(r'(<div>数据周期: 2026-07-01 至 )2026-07-27',r'\g<1>2026-07-28')
text=text.replace('7月累计27天','7月累计28天').replace('7月24小时时段数据（7.1-7.27累计）','7月24小时时段数据（7.1-7.28累计）').replace('7月人员排行(7.1-7.27累计)','7月人员排行(7.1-7.28累计)')
HTML.write_text(text,encoding='utf-8')
dash={'through':'7.28','cum_actual':cum_actual,'cum_trans':cum_trans,'people_total':int(people_total),'eff':eff,'cum_viol':cum_viol,'cum_queue':cum_queue,'daily':{d:{'actual':a,'trans':t,'queue':q,'viol':v} for d,a,t,q,v in zip(labels,audit,trans,queues,viol)},'person_rank':pr,'apps':dict(apps),'hourly':hr,'zt_volumes':zv,'zt_queues':zq}
(ROOT/'dashboard_7.28_summary.json').write_text(json.dumps(dash,ensure_ascii=False,indent=2),encoding='utf-8')
print(json.dumps({'reports':reports,'workbook_before':len(old),'workbook_after':len(post),'added':added,'dashboard':dash},ensure_ascii=False,indent=2))
